import pyodbc
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

def get_connection(driver='SQL Server Native Client 11.0', server='localhost', database='NexttLoja', username='sa', password=None, trusted_connection='yes'):
    string_connection = f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password};Trusted_Connection={trusted_connection}"
    connection = pyodbc.connect(string_connection)
    cursor = connection.cursor()
    return connection, cursor

def obter_dados_necessarios():
    connection, cursor = get_connection()

    consultas = {
        "sec_codigo": "SELECT sec_codigo FROM tb_secao",
        "esp_codigo": "SELECT CAST(esp_codigo AS VARCHAR) AS descricao_completa FROM tb_especie;",
        "mar_codigo": "SELECT mar_codigo FROM tb_marca",
        "usu_codigo": "SELECT usu_codigo FROM tb_usuario WHERE usu_codigo <> 1 and usu_codigo <> 2",
        "und_codigo": "SELECT und_codigo FROM tb_unidade",
        "etq_codigo": "SELECT etq_codigo FROM tb_etiqueta",
        "secao_completa": "SELECT CONCAT(sec_codigo, ' - ', sec_descricao) AS descricao_completa FROM tb_secao",
        "especie_completa": "SELECT CAST(esp_codigo AS VARCHAR) + ' - ' + LTRIM(SUBSTRING(esp_descricao, PATINDEX('%[A-Z]%', esp_descricao), LEN(esp_descricao))) AS descricao_completa FROM tb_especie;",
        "marca_completa": "SELECT CONCAT(mar_codigo, ' - ', mar_descricao) AS descricao_completa FROM tb_marca;",
        "comprador_completo": "SELECT CONCAT(usu_codigo, ' - ', usu_nome) AS descricao_completa FROM tb_usuario WHERE set_codigo IS NULL and usu_codigo <> 1 and usu_codigo <> 2",
        "unidade_completa": "SELECT CONCAT(und_codigo, ' - ', und_descricao) AS descricao_completa FROM tb_unidade ",
        "etiqueta_completa": "SELECT CONCAT(etq_codigo, ' - ', etq_descricao) AS descricao_completa FROM tb_etiqueta ",
        "empresa_nome": "SELECT emp_descricao FROM tb_empresa",
        "clf_codigo": "SELECT MIN(clf_codigo) AS clf_codigo FROM tb_classificacao_fiscal WHERE clf_ativo = 1 GROUP BY clf_descricao ORDER BY clf_codigo ASC",
        "classificacao_completa": "SELECT CONCAT(MIN(clf_codigo_fiscal), ' - ', clf_descricao) AS descricao_completa FROM tb_classificacao_fiscal WHERE clf_ativo = 1 GROUP BY clf_descricao ORDER BY descricao_completa ASC"
    }

    resultados = {}

    for chave, query in consultas.items():
        cursor.execute(query)
        resultados[chave] = [row[0] for row in cursor.fetchall()]
    
    connection.close()
    return resultados

def carregar_planilha(caminho_arquivo):
    try:
        wb = openpyxl.load_workbook(caminho_arquivo)
        return wb
    except FileNotFoundError:
        print("Arquivo não encontrado.")
        return None
    except Exception as e:
        print(f"Erro ao carregar o arquivo: {e}")
        return None

def preencher_planilha(dados, caminho_arquivo):
    wb = carregar_planilha(caminho_arquivo)
    if not wb:
        print("Não foi possível carregar o arquivo. Verifique se o arquivo existe e tente novamente.")
        return

    nome_aba_dados = "Dados Consolidados"
    if nome_aba_dados in wb.sheetnames:
        aba_dados = wb[nome_aba_dados]
    else:
        aba_dados = wb.create_sheet(title=nome_aba_dados)

    mapeamento_colunas = {
        "sec_codigo": "R",
        "esp_codigo": "S",
        "mar_codigo": "T",
        "usu_codigo": "U",
        "und_codigo": "V",
        "etq_codigo": "W",
        "clf_codigo": "X",
        "secao_completa": "A",
        "especie_completa": "B",
        "marca_completa": "E",
        "comprador_completo": "H",
        "unidade_completa": "J",
        "classificacao_completa": "K",
        "etiqueta_completa": "P"
        
    }

    for coluna in mapeamento_colunas.values():
        for linha in range(2, aba_dados.max_row + 1):
            aba_dados[f"{coluna}{linha}"].value = None  

    for chave, lista in dados.items():
        coluna = mapeamento_colunas.get(chave)
        if coluna is not None:
            for j, valor in enumerate(lista, start=2):
                aba_dados[f"{coluna}{j}"] = valor

    nome_aba_planilha = "Cadastro de Produtos"
    if nome_aba_planilha in wb.sheetnames:
        aba_planilha = wb[nome_aba_planilha]
    else:
        aba_planilha = wb.create_sheet(title=nome_aba_planilha)

    empresa_nome = dados.get("empresa_nome")[0]
    aba_planilha['A2'] = f"Cadastro de Produtos {empresa_nome}"

    wb.save(caminho_arquivo)

    def adicionar_validacao_coluna(coluna, dados_coluna, primeira_linha=7):
        primeira_linha_dados = 2
        ultima_linha_dados = primeira_linha_dados + len(dados_coluna) - 1
        referencia_validacao = f"'{nome_aba_dados}'!${coluna}${primeira_linha_dados}:${coluna}${ultima_linha_dados}"
        
        dv = DataValidation(type="list", formula1=referencia_validacao, showDropDown=False)
        dv.error = "Por favor, selecione um valor da lista."
        dv.errorTitle = "Valor Inválido"
        dv.showErrorMessage = True
        
        aba_planilha.add_data_validation(dv)

        for i in range(primeira_linha, primeira_linha + max_linhas):
            aba_planilha[f"{coluna}{i}"].value = None  
            dv.add(aba_planilha[f"{coluna}{i}"])  

    max_linhas = max(len(lista) for lista in dados.values()) + 10  

    adicionar_validacao_coluna("A", dados["secao_completa"])  
    adicionar_validacao_coluna("B", dados["especie_completa"])  
    adicionar_validacao_coluna("E", dados["marca_completa"])  
    adicionar_validacao_coluna("H", dados["comprador_completo"])  
    adicionar_validacao_coluna("J", dados["unidade_completa"])  
    adicionar_validacao_coluna("K", dados["classificacao_completa"])  
    adicionar_validacao_coluna("P", dados["etiqueta_completa"])  
    adicionar_validacao_coluna("R", dados["sec_codigo"])  
    adicionar_validacao_coluna("S", dados["esp_codigo"])  
    adicionar_validacao_coluna("T", dados["mar_codigo"])  
    adicionar_validacao_coluna("U", dados["usu_codigo"])  
    adicionar_validacao_coluna("V", dados["und_codigo"])
    adicionar_validacao_coluna("W", dados["etq_codigo"])  
    adicionar_validacao_coluna("X", dados["clf_codigo"])  

    wb.save(caminho_arquivo)
    wb.close()

caminho_arquivo = 'Cadastros Auto Nextt teste.xlsx'
dados = obter_dados_necessarios()

preencher_planilha(dados, caminho_arquivo)

print(f"Dados preenchidos e validação de dados adicionada.")
