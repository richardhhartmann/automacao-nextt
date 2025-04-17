import pandas as pd
import pyodbc
import tkinter as tk
import os
import json
import sys
import openpyxl
import time
from tkinter import filedialog
from datetime import datetime
from openpyxl.utils import get_column_letter, column_index_from_string
from contextlib import contextmanager

BATCH_SIZE = 100
LINHA_CABECALHO = 3
COLUNA_INICIAL_ADICIONAIS = 'Z'
COLUNAS_VARIACOES = [('R', 'V'), ('S', 'W'), ('T', 'X'), ('U', 'Y')]
DEBUG = True

def debug_log(message):
    if DEBUG:
        print(f"[DEBUG] {message}")

@contextmanager
def get_db_connection(file_name='conexao_temp.txt'):
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(script_dir, file_name)
        
        with open(file_path, 'r') as f:
            config = json.load(f)

        driver = config.get('driver', None)
        server = config.get('server', None)
        database = config.get('database', None)
        username = config.get('username', None)
        password = config.get('password', None)
        trusted_connection = config.get('trusted_connection', None)

        if trusted_connection and trusted_connection.lower() == 'yes':
            conn_str = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};Trusted_Connection={trusted_connection}"
        else:
            conn_str = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={username};PWD={password};"

        connection = pyodbc.connect(conn_str)
        connection.autocommit = False
        cursor = connection.cursor()
        debug_log("Conexão com o banco estabelecida")
        
        yield connection, cursor
        
    except Exception as e:
        debug_log(f"ERRO na conexão: {e}")
        sys.exit(1)
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

def selecionar_arquivo():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Arquivos Excel", "*.xlsx;*.xls;*.xlsm")]
    )

def get_colunas_adicionais(ws, linha_cabecalho=LINHA_CABECALHO, coluna_inicial=COLUNA_INICIAL_ADICIONAIS):
    colunas = []
    current_col_idx = column_index_from_string(coluna_inicial)

    while True:
        col_letter = get_column_letter(current_col_idx)
        valor = ws[f'{col_letter}{linha_cabecalho}'].value
        if pd.isna(valor) or not valor:
            break
        colunas.append(col_letter)
        current_col_idx += 1

    return colunas

def trata_valor(valor, tipo=int):
    if pd.isna(valor) or str(valor).strip().upper() in ['#N/D', '#N/A', 'N/D', '']:
        return None
    try:
        return tipo(valor)
    except (ValueError, TypeError):
        return None

def processa_produto(ws, linha_excel, df, x):
    if df.iloc[x, 62] != "OK":
        return None

    secao = trata_valor(df.iloc[x, 54])
    especie = trata_valor(df.iloc[x, 55])
    descricao = str(df.iloc[x, 2])[:50] if pd.notna(df.iloc[x, 2]) else None
    descricao_reduzida = str(df.iloc[x, 3])[:50] if pd.notna(df.iloc[x, 3]) else None
    marca = trata_valor(df.iloc[x, 56])
    comprador = trata_valor(df.iloc[x, 57])
    und_codigo = trata_valor(df.iloc[x, 58])
    classificacao = trata_valor(df.iloc[x, 59])
    origem = trata_valor(df.iloc[x, 60])
    etiqueta = trata_valor(df.iloc[x, 61])

    referencia = (str(int(df.iloc[x, 5])) if isinstance(df.iloc[x, 5], float) and df.iloc[x, 5].is_integer() else str(df.iloc[x, 5])) if pd.notna(df.iloc[x, 5]) else None
    cod_original = str(df.iloc[x, 6]) if pd.notna(df.iloc[x, 6]) else ''
    
    ativo = 1
    venda = float(str(df.iloc[x, 12]).replace(',', '.')) if pd.notna(df.iloc[x, 12]) else None
    icms = float(df.iloc[x, 13]) if pd.notna(df.iloc[x, 13]) else None
    ipi = float(df.iloc[x, 14]) if pd.notna(df.iloc[x, 14]) else None
    ipr_codigo_barra = trata_valor(df.iloc[x, 16])
    
    data = datetime.now()
    
    atributos_adicionais = []
    colunas_adicionais = get_colunas_adicionais(ws)
    
    for idx, col in enumerate(colunas_adicionais, start=3):
        valor = ws[f'{col}{linha_excel}'].value
        if pd.notna(valor):
            atributos_adicionais.append((secao, especie, None, idx, str(valor), None))

    variacoes = []
    for i, (col_cor, col_tamanho) in enumerate(COLUNAS_VARIACOES, start=1):
        cor = ws[f'{col_cor}{linha_excel}'].value
        tamanho = ws[f'{col_tamanho}{linha_excel}'].value

        if pd.notna(cor) and pd.notna(tamanho):
            variacoes.append({
                'ipr_codigo': i,
                'ipr_codigo_barra': ipr_codigo_barra,
                'cor': str(cor),
                'tamanho': str(tamanho)
            })

    return {
        'secao': secao,
        'especie': especie,
        'descricao': descricao,
        'descricao_reduzida': descricao_reduzida,
        'marca': marca,
        'comprador': comprador,
        'und_codigo': und_codigo,
        'classificacao': classificacao,
        'origem': origem,
        'etiqueta': etiqueta,
        'referencia': referencia,
        'cod_original': cod_original,
        'ativo': ativo,
        'venda': venda,
        'icms': icms,
        'ipi': ipi,
        'data': data,
        'atributos_adicionais': atributos_adicionais,
        'variacoes': variacoes,
        'linha_excel': linha_excel
    }

def verificar_duplicados(cursor, referencias_marcas):
    if not referencias_marcas:
        return set()
    
    cursor.execute("CREATE TABLE #TempDuplicados (Referencia NVARCHAR(255), Marca INT)")
    cursor.executemany("INSERT INTO #TempDuplicados (Referencia, Marca) VALUES (?, ?)", referencias_marcas)
    
    cursor.execute("""
        SELECT t.Referencia, t.Marca 
        FROM #TempDuplicados t
        INNER JOIN tb_produto p ON p.prd_referencia_fornec = t.Referencia AND p.mar_codigo = t.Marca
    """)
    
    duplicados = set((row.Referencia, row.Marca) for row in cursor.fetchall())
    cursor.execute("DROP TABLE #TempDuplicados")
    
    debug_log(f"Encontrados {len(duplicados)} produtos duplicados")
    return duplicados

def cadastrar_produto():
    debug_log("Iniciando processo de cadastro")
    start_total = time.time()
    
    caminho_arquivo = selecionar_arquivo()
    if not caminho_arquivo:
        return

    try:
        debug_log("Carregando arquivo Excel")
        wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        ws = wb["Cadastro de Produtos"]
        
        df = pd.read_excel(caminho_arquivo, sheet_name="Cadastro de Produtos", skiprows=6, header=None)
        df = df.dropna(how='all')
        
        df.columns = [
            "secao", "especie", "descricao", "descricao_reduzida", "marca", "referencia",
            "cod_original", "comprador", "ativo", "unidade", "classificacao", "origem",
            "venda", "icms", "ipi", "etiqueta", "coluna17", "coluna18", "coluna19", "coluna20",
            "coluna21", "coluna22", "coluna23", "coluna24", "coluna25", "coluna26", "coluna27",
            "coluna28", "coluna29", "coluna30", "coluna31"
        ] + [f"coluna{i}" for i in range(32, len(df.columns) + 1)]

        colunas_numericas = ['secao', 'especie', 'cod_original', 'classificacao', 'venda', 'origem', 
                            'icms', 'ipi', 'etiqueta', 'comprador', 'coluna25', 'coluna26', 'coluna27',
                            'coluna28', 'coluna29', 'coluna30', 'coluna31', 'coluna54', 'coluna55',
                            'coluna56', 'coluna57', 'coluna58', 'coluna59', 'coluna60', 'coluna61',
                            'coluna62']
        
        for col in colunas_numericas:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype('int16')

        produtos = []
        for x in range(len(df)):
            linha_excel = x + 7
            produto = processa_produto(ws, linha_excel, df, x)
            if produto:
                produtos.append(produto)
        
        debug_log(f"Total de produtos a processar: {len(produtos)}")

        with get_db_connection('conexao_temp.txt') as (connection, cursor):
            duplicata = 0
            produtos_inseridos = 0
            variacoes_inseridas = 0
            
            referencias_marcas = [(p['referencia'], p['marca']) for p in produtos if p['referencia'] and p['marca']]
            duplicados = verificar_duplicados(cursor, referencias_marcas)
            
            for i, produto in enumerate(produtos, 1):
                if (produto['referencia'], produto['marca']) in duplicados:
                    duplicata += 1
                    continue
                
                cursor.execute("SELECT MAX(prd_codigo) FROM tb_produto WHERE sec_codigo = ? AND esp_codigo = ?", produto['secao'], produto['especie'])
                resultado = cursor.fetchone()
                prd_codigo = (resultado[0] or 0) + 1
                
                parametros_produto = (
                    produto['secao'], produto['especie'], prd_codigo, produto['descricao'], 
                    produto['descricao_reduzida'], produto['marca'], produto['data'], None, None, None,
                    0, 0, 0, produto['cod_original'], produto['ativo'], 0, None, produto['referencia'],
                    None, produto['classificacao'], produto['comprador'], produto['und_codigo'], 
                    produto['venda'], produto['origem'], produto['icms'], produto['ipi'], 
                    produto['etiqueta'], None, None, None, 1, None, None, None, None, None
                )
                
                cursor.execute("""
                    INSERT INTO tb_produto
                    (sec_codigo, esp_codigo, prd_codigo, prd_descricao, prd_descricao_reduzida,
                    mar_codigo, prd_data_cadastro, prd_unidade, prd_data_ultima_compra,
                    prd_data_ultima_entrega, prd_custo_medio, prd_preco_medio, prd_aliquota_icms,
                    prd_codigo_original, prd_ativo, prd_ultimo_custo, prd_arquivo_foto,
                    prd_referencia_fornec, prd_tipo_tributacao, clf_codigo, usu_codigo_comprador,
                    und_codigo, prd_valor_venda, prd_origem, prd_percentual_icms, prd_percentual_ipi,
                    etq_codigo_padrao, usu_codigo_cadastro, sec_codigo_r, esp_codigo_r,
                    prd_permite_comprar, prd_valor_unidade_conversao, udc_codigo, und_codigo_conversao,
                    prd_iat, prd_ippt)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
                            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, parametros_produto)
                
                for atributo in produto['atributos_adicionais']:
                    sec, esp, _, tpa, desc, _ = atributo

                    cursor.execute("""
                        SELECT COUNT(*) 
                        FROM INFORMATION_SCHEMA.COLUMNS 
                        WHERE TABLE_NAME = 'tb_atributo_produto' 
                        AND COLUMN_NAME = 'prr_codigo'
                    """)
                    has_prr_codigo = cursor.fetchone()[0] > 0
                    
                    if has_prr_codigo:
                        try:
                            cursor.execute("""
                                INSERT INTO tb_atributo_produto 
                                (sec_codigo, esp_codigo, prd_codigo, tpa_codigo, apr_descricao, prr_codigo)
                                VALUES (?, ?, ?, ?, ?, NULL)
                            """, sec, esp, prd_codigo, tpa, desc)
                        except pyodbc.IntegrityError:
                            pass
                    else:
                        cursor.execute("""
                            INSERT INTO tb_atributo_produto 
                            (sec_codigo, esp_codigo, prd_codigo, tpa_codigo, apr_descricao)
                            VALUES (?, ?, ?, ?, ?)
                        """, (sec, esp, prd_codigo, tpa, desc))
                
                cursor.execute("""SELECT COUNT(*) 
                    FROM INFORMATION_SCHEMA.COLUMNS 
                    WHERE TABLE_NAME = 'tb_atributo_produto' 
                    AND COLUMN_NAME = 'ipr_gtin'""")
                
                has_ipr_gtin = cursor.fetchone()[0] > 0

                for variacao in produto['variacoes']:
                    if has_ipr_gtin:
                        cursor.execute("""
                            INSERT INTO tb_item_produto 
                            (sec_codigo, esp_codigo, prd_codigo, ipr_codigo, ipr_codigo_barra, ipr_preco_promocional, ipr_gtin)
                            VALUES (?, ?, ?, ?, ?, 0, NULL)
                        """, produto['secao'], produto['especie'], prd_codigo, variacao['ipr_codigo'], variacao['ipr_codigo_barra'])
                    else:
                        cursor.execute("""
                            INSERT INTO tb_item_produto 
                            (sec_codigo, esp_codigo, prd_codigo, ipr_codigo, ipr_codigo_barra, ipr_preco_promocional)
                            VALUES (?, ?, ?, ?, ?, 0)
                        """, produto['secao'], produto['especie'], prd_codigo, variacao['ipr_codigo'], variacao['ipr_codigo_barra'])

                    
                    cursor.execute("""
                        INSERT INTO tb_atributo_item_produto 
                        (sec_codigo, esp_codigo, prd_codigo, ipr_codigo,
                        tpa_codigo, aip_descricao, aip_ordem, aip_descricao_fornec)
                        VALUES (?, ?, ?, ?, 1, ?, ?, NULL)
                    """, produto['secao'], produto['especie'], prd_codigo, variacao['ipr_codigo'], variacao['cor'], variacao['ipr_codigo'])
                    
                    cursor.execute("""
                        INSERT INTO tb_atributo_item_produto 
                        (sec_codigo, esp_codigo, prd_codigo, ipr_codigo,
                        tpa_codigo, aip_descricao, aip_ordem, aip_descricao_fornec)
                        VALUES (?, ?, ?, ?, 2, ?, ?, NULL)
                    """, produto['secao'], produto['especie'], prd_codigo, variacao['ipr_codigo'], variacao['tamanho'], variacao['ipr_codigo'])
                    
                    variacoes_inseridas += 1
                
                produtos_inseridos += 1
                
                if i % BATCH_SIZE == 0:
                    connection.commit()
                    debug_log(f"Lote {i//BATCH_SIZE} commitado ({i} produtos processados)")
            
            connection.commit()
            
            tempo_total = time.time() - start_total
            debug_log(f"Processo concluído em {tempo_total:.2f} segundos")
            print("\nResumo:")
            print(f"- Total de produtos processados: {len(produtos)}")
            print(f"- Produtos inseridos: {produtos_inseridos}")
            print(f"- Variações inseridas: {variacoes_inseridas}")
            print(f"- Produtos duplicados (ignorados): {duplicata}")
            print(f"- Tempo médio por produto: {tempo_total/max(1, produtos_inseridos):.2f}s")
            
    except Exception as e:
        debug_log(f"ERRO: {str(e)}")
        if 'connection' in locals():
            connection.rollback()
    finally:
        if 'wb' in locals():
            wb.close()

if __name__ == "__main__":
    cadastrar_produto()