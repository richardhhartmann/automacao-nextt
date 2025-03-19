import pyodbc
import openpyxl
import win32com.client as win32  # Para executar a macro
from openpyxl.worksheet.datavalidation import DataValidation
import time
from tqdm import tqdm  # Biblioteca para barra de progresso

def get_connection(driver='SQL Server Native Client 11.0', server='localhost', database='NexttLoja', username='sa', password=None, trusted_connection='yes'):
    string_connection = f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password};Trusted_Connection={trusted_connection}"
    connection = pyodbc.connect(string_connection)
    cursor = connection.cursor()
    return connection, cursor

def obter_dados_necessarios():
    inicio = time.time()  # Marca o tempo inicial
    connection, cursor = get_connection()

    consultas = {
        "sec_codigo": "SELECT sec_codigo FROM tb_secao",
        "esp_codigo": "SELECT CAST(esp_codigo AS VARCHAR) AS descricao_completa FROM tb_especie;",
        "mar_codigo": "SELECT mar_codigo FROM tb_marca",
        "usu_codigo": "SELECT usu_codigo FROM tb_usuario WHERE usu_codigo <> 1 and usu_codigo <> 2",
        "und_codigo": "SELECT und_codigo FROM tb_unidade",
        "etq_codigo": "SELECT etq_codigo FROM tb_etiqueta",
        "secao_completa": "SELECT CONCAT(sec_codigo, ' - ', sec_descricao) AS descricao_completa FROM tb_secao",
        "especie_completa": "SELECT CAST(esp_codigo AS VARCHAR) + ' - ' + LTRIM(SUBSTRING(esp_descricao, PATINDEX('%[A-Z]%', esp_descricao), LEN(esp_descricao))) AS descricao_completa FROM tb_especie;",
        "marca_completa": "SELECT CONCAT(mar_codigo, ' - ', mar_descricao) AS descricao_completa FROM tb_marca;",
        "comprador_completo": "SELECT CONCAT(usu_codigo, ' - ', usu_nome) AS descricao_completa FROM tb_usuario WHERE set_codigo IS NULL and usu_codigo <> 1 and usu_codigo <> 2",
        "unidade_completa": "SELECT CONCAT(und_codigo, ' - ', und_descricao) AS descricao_completa FROM tb_unidade ",
        "etiqueta_completa": "SELECT CONCAT(etq_codigo, ' - ', etq_descricao) AS descricao_completa FROM tb_etiqueta ",
        "empresa_nome": "SELECT emp_descricao FROM tb_empresa",
        "clf_codigo": "SELECT MIN(clf_codigo) AS clf_codigo FROM tb_classificacao_fiscal WHERE clf_ativo = 1 GROUP BY clf_descricao ORDER BY clf_codigo ASC",
        "classificacao_completa": "SELECT CONCAT(MIN(clf_codigo_fiscal), ' - ', clf_descricao) AS descricao_completa FROM tb_classificacao_fiscal WHERE clf_ativo = 1 GROUP BY clf_descricao ORDER BY descricao_completa ASC"
    }

    resultados = {}

    print("Buscando dados do banco de dados...")
    for chave, query in tqdm(consultas.items(), desc="Executando queries"):  # Barra de progresso
        cursor.execute(query)
        resultados[chave] = [row[0] for row in cursor.fetchall()]
    
    connection.close()

    print(f"Tempo total para obter dados: {time.time() - inicio:.2f} segundos")
    return resultados

def preencher_planilha(dados, caminho_arquivo):
    inicio = time.time()  # Tempo inicial
    wb = openpyxl.load_workbook(caminho_arquivo)
    
    nome_aba_dados = "Dados Consolidados"
    if nome_aba_dados in wb.sheetnames:
        aba_dados = wb[nome_aba_dados]
    else:
        aba_dados = wb.create_sheet(title=nome_aba_dados)

    mapeamento_colunas = {
        "sec_codigo": "R",
        "esp_codigo": "S",
        "mar_codigo": "T",
        "usu_codigo": "U",
        "und_codigo": "V",
        "etq_codigo": "W",
        "clf_codigo": "X",
        "secao_completa": "A",
        "especie_completa": "B",
        "marca_completa": "E",
        "comprador_completo": "H",
        "unidade_completa": "J",
        "classificacao_completa": "K",
        "etiqueta_completa": "P"
    }

    print("Limpando planilha e preenchendo novos dados...")
    for coluna in tqdm(mapeamento_colunas.values(), desc="Limpando planilha"):
        for linha in range(2, aba_dados.max_row + 1):
            aba_dados[f"{coluna}{linha}"].value = None  

    for chave, lista in tqdm(dados.items(), desc="Preenchendo planilha"):
        coluna = mapeamento_colunas.get(chave)
        if coluna is not None:
            for j, valor in enumerate(lista, start=2):
                aba_dados[f"{coluna}{j}"] = valor

    nome_aba_planilha = "Cadastro de Produtos"
    if nome_aba_planilha in wb.sheetnames:
        aba_planilha = wb[nome_aba_planilha]
    else:
        aba_planilha = wb.create_sheet(title=nome_aba_planilha)

    empresa_nome = dados.get("empresa_nome")[0]
    aba_planilha['A2'] = f"Cadastro de Produtos {empresa_nome}"
    print(f"Definindo nome da empresa '{empresa_nome}' em: '{aba_planilha.title}'")

    wb.save(caminho_arquivo)

    def adicionar_validacao_coluna(coluna, dados_coluna, primeira_linha=7):
        primeira_linha_dados = 2
        ultima_linha_dados = primeira_linha_dados + len(dados_coluna) - 1
        referencia_validacao = f"'{nome_aba_dados}'!${coluna}${primeira_linha_dados}:${coluna}${ultima_linha_dados}"
        
        dv = DataValidation(type="list", formula1=referencia_validacao, showDropDown=False)
        dv.error = "Por favor, selecione um valor da lista."
        dv.errorTitle = "Valor Inválido"
        dv.showErrorMessage = True
        
        aba_planilha.add_data_validation(dv)

        for i in range(primeira_linha, primeira_linha + max_linhas):
            aba_planilha[f"{coluna}{i}"].value = None  
            dv.add(aba_planilha[f"{coluna}{i}"])  

    # Atualizar a validação de dados na coluna B (a partir da linha 7)
    print("Atualizando validação de dados na coluna B...")
    for i in range(7, aba_planilha.max_row + 1):
        formula = f'=INDIRETO("\'Dados Consolidados\'!SeçãoCompleta" & Y{i})'
        
        dv = DataValidation(type="list", formula1=formula, showDropDown=False)
        dv.error = "Por favor, selecione um valor da lista."
        dv.errorTitle = "Valor Inválido"
        dv.showErrorMessage = True
        
        aba_planilha.add_data_validation(dv)
        dv.add(aba_planilha[f"B{i}"])  # Aplica a validação na célula da linha i

    max_linhas = max(len(lista) for lista in dados.values()) + 10  

    colunas_com_validacao = ["A", "B", "E", "H", "J", "K", "P"]

    print("Adicionando validação de dados...")    
    for chave, coluna in tqdm(mapeamento_colunas.items(), desc="Validação"):
        if coluna in colunas_com_validacao and chave in dados:  # Aplicar apenas para as colunas desejadas
            adicionar_validacao_coluna(coluna, dados[chave])

    wb.save(caminho_arquivo)
    wb.close()

    tempo_total = time.time() - inicio

    if tempo_total > 60:
        minutos = int(tempo_total // 60)
        segundos = tempo_total % 60
        print(f"Tempo total para preencher planilha: {minutos} minutos e {segundos:.0f} segundos")
    else:
        print(f"Tempo total para preencher planilha: {tempo_total:.0f} segundos")

def adicionar_macro_vba(caminho_arquivo):
    # Inicializa o Excel e abre o arquivo
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = True  # Não mostrar a janela do Excel

    wb = excel.Workbooks.Open(caminho_arquivo)
    vb_module = wb.VBProject.VBComponents.Add(1)  # Adiciona um módulo de código VBA

    # Código VBA da macro
    vba_code = """
    Sub CriarIntervalosNomeadosB()
        Dim ws As Worksheet
        Dim ultimaLinha As Long
        Dim inicio As Long
        Dim i As Long
        Dim secao As Integer
        Dim nomeSecao As String
        Dim rng As Range

        ' Definir a planilha
        Set ws = ThisWorkbook.Sheets("Dados Consolidados")

        ' Encontrar a última linha preenchida na coluna B
        ultimaLinha = ws.Cells(ws.Rows.Count, 2).End(xlUp).Row

        ' Inicializar variáveis
        secao = 1
        inicio = 2 ' Começar na linha 2

        ' Percorrer a coluna B
        For i = 2 To ultimaLinha
            If Left(ws.Cells(i, 2).Value, 2) = "1 " Then
                ' Criar um intervalo nomeado para a seção anterior (se existir)
                If inicio < i Then
                    nomeSecao = "SeçãoCompleta" & secao
                    Set rng = ws.Range("B" & inicio & ":B" & (i - 1))
                    ws.Names.Add Name:=nomeSecao, RefersTo:=rng
                    secao = secao + 1
                End If
                ' Atualizar o início para a nova seção
                inicio = i
            End If
        Next i

        ' Criar o último intervalo
        If inicio <= ultimaLinha Then
            nomeSecao = "SeçãoCompleta" & secao
            Set rng = ws.Range("B" & inicio & ":B" & ultimaLinha)
            ws.Names.Add Name:=nomeSecao, RefersTo:=rng
        End If

        MsgBox "Intervalos nomeados criados com sucesso!", vbInformation
    End Sub
    """

    # Adiciona o código VBA no módulo
    vb_module.CodeModule.AddFromString(vba_code)
    wb.Save()  # Salva o arquivo após adicionar o código
    wb.Close()
    excel.Quit()

def executar_macro(caminho_arquivo):
    # Inicializa o Excel para executar a macro
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False  # Não mostrar o Excel

    wb = excel.Workbooks.Open(caminho_arquivo)
    
    # Executa a macro
    excel.Application.Run("CriarIntervalosNomeadosB")

    wb.Save()  # Salva o arquivo após a execução da macro
    wb.Close()
    excel.Quit()

# Executando o script
caminho_arquivo = 'Cadastros Auto Nextt limpa.xlsx'
dados = obter_dados_necessarios()
preencher_planilha(dados, caminho_arquivo)

# Agora adiciona a macro VBA ao arquivo
adicionar_macro_vba(caminho_arquivo)

# Executa a macro após o preenchimento da planilha
executar_macro(caminho_arquivo)

print("Dados preenchidos e macro executada com sucesso.")
