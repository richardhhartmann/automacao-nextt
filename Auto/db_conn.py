import os
import pyodbc
import openpyxl
import shutil
import win32com.client as win32
import winreg as reg
import time
from openpyxl.worksheet.datavalidation import DataValidation
from tqdm import tqdm

def get_connection(driver='SQL Server Native Client 11.0', server='localhost', database='NexttLoja', username='sa', password=None, trusted_connection='yes'):
    string_connection = f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password};Trusted_Connection={trusted_connection}"
    connection = pyodbc.connect(string_connection)
    cursor = connection.cursor()
    return connection, cursor

def dados_necessarios():
    print(f"Adicionando informações à planilha: {caminho_arquivo}")
    inicio = time.time()
    connection, cursor = get_connection()

    consultas = {
        "sec_codigo": "SELECT sec_codigo FROM tb_secao",
        "esp_codigo": "SELECT CAST(esp_codigo AS VARCHAR) AS descricao_completa FROM tb_especie;",
        "mar_codigo": "SELECT mar_codigo FROM tb_marca",
        "usu_codigo": "SELECT usu_codigo FROM tb_usuario WHERE usu_codigo <> 1 and usu_codigo <> 2",
        "und_codigo": "SELECT und_codigo FROM tb_unidade",
        "etq_codigo": "SELECT etq_codigo FROM tb_etiqueta",
        "secao_completa": "SELECT CONCAT(sec_codigo, ' - ', sec_descricao) AS descricao_completa FROM tb_secao",
        "especie_completa": "SELECT CAST(esp_codigo AS VARCHAR) + ' - ' + LTRIM(SUBSTRING(esp_descricao, PATINDEX('%[A-Z]%', esp_descricao), LEN(esp_descricao))) AS descricao_completa FROM tb_especie;",
        "marca_completa": "SELECT CONCAT(mar_codigo, ' - ', mar_descricao) AS descricao_completa FROM tb_marca;",
        "comprador_completo": "SELECT CONCAT(usu_codigo, ' - ', usu_nome) AS descricao_completa FROM tb_usuario WHERE set_codigo IS NULL and usu_codigo <> 1 and usu_codigo <> 2",
        "unidade_completa": "SELECT und_descricao from tb_unidade ",
        "etiqueta_completa": "SELECT CONCAT(etq_codigo, ' - ', etq_descricao) AS descricao_completa FROM tb_etiqueta ",
        "empresa_nome": "SELECT emp_descricao FROM tb_empresa",
        "clf_codigo": "SELECT MIN(clf_codigo) AS clf_codigo FROM tb_classificacao_fiscal WHERE clf_ativo = 1 GROUP BY clf_descricao ORDER BY clf_codigo ASC",
        "classificacao_completa": "SELECT CONCAT(MIN(clf_codigo_fiscal), ' - ', clf_descricao) AS descricao_completa FROM tb_classificacao_fiscal WHERE clf_ativo = 1 GROUP BY clf_descricao ORDER BY descricao_completa ASC",
        "seg_codigo": "SELECT seg_codigo FROM tb_segmento",
        "segmento_completo": "SELECT seg_descricao FROM tb_segmento",
        "marca_descricao": "SELECT mar_descricao from tb_marca"
    }

    resultados = {}

    print("Buscando dados do banco de dados...")
    for chave, query in tqdm(consultas.items(), desc="Executando queries"):
        cursor.execute(query)
        resultados[chave] = [row[0] for row in cursor.fetchall()]
    
    connection.close()

    print(f"Tempo total para obter dados: {time.time() - inicio:.2f} segundos\n")
    return resultados

def preencher_planilha(dados, caminho_arquivo):
    # Cria uma cópia da planilha original, renomeando para 'Cadastros Auto Nextt.xlsx'
    caminho_arquivo_novo = caminho_arquivo.replace("Cadastros Auto Nextt limpa", "Cadastros Auto Nextt")
    shutil.copy(caminho_arquivo, caminho_arquivo_novo)

    # Agora, vamos preencher a nova planilha copiada
    inicio = time.time()
    wb = openpyxl.load_workbook(caminho_arquivo_novo)
    
    nome_aba_dados = "Dados Consolidados"
    if nome_aba_dados in wb.sheetnames:
        aba_dados = wb[nome_aba_dados]
    else:
        aba_dados = wb.create_sheet(title=nome_aba_dados)

    mapeamento_colunas = {
        "sec_codigo": "R",
        "esp_codigo": "S",
        "mar_codigo": "T",
        "usu_codigo": "U",
        "und_codigo": "V",
        "etq_codigo": "W",
        "clf_codigo": "X",
        "secao_completa": "A",
        "especie_completa": "B",
        "marca_completa": "E",
        "comprador_completo": "H",
        "unidade_completa": "J",
        "classificacao_completa": "K",
        "etiqueta_completa": "P",
        "segmento_completo": "AR",
        "seg_codigo": "AS",
        "marca_descricao": "AT"
    }

    print("Limpando planilha e preenchendo novos dados...")
    for coluna in tqdm(mapeamento_colunas.values(), desc="Limpando planilha"):
        for linha in range(2, aba_dados.max_row + 1):
            aba_dados[f"{coluna}{linha}"].value = None  

    for chave, lista in tqdm(dados.items(), desc="Preenchendo planilha"):
        coluna = mapeamento_colunas.get(chave)
        if coluna is not None:
            for j, valor in enumerate(lista, start=1):
                aba_dados[f"{coluna}{j}"] = valor

    nome_aba_planilha = "Cadastro de Produtos"
    if nome_aba_planilha in wb.sheetnames:
        aba_planilha = wb[nome_aba_planilha]
    else:
        aba_planilha = wb.create_sheet(title=nome_aba_planilha)

    empresa_nome = dados.get("empresa_nome")[0]
    aba_planilha['A2'] = f"Cadastro de Produtos {empresa_nome}"
    print(f"\nDefinindo nome da empresa '{empresa_nome}' em: '{aba_planilha.title}'\n")

    wb.save(caminho_arquivo_novo)

    # Função de adicionar validação
    def adicionar_validacao(aba, intervalo_celulas, referencia_dados):

        dv = DataValidation(type="list", formula1=referencia_dados, showDropDown=False)
        dv.error = "Por favor, selecione um valor da lista."
        dv.errorTitle = "Valor Inválido"
        dv.showErrorMessage = True

        # Aplicar validação às células do intervalo
        aba.add_data_validation(dv)
        for linha in aba[intervalo_celulas]:
            for celula in linha:
                dv.add(celula)

    # Atualizando a validação de dados na coluna B com a fórmula dinâmica
    print("Atualizando validação de dados na coluna B...")
    for i in range(7, aba_planilha.max_row + 1):
        # A fórmula agora usa a referência indireta à célula da coluna Y para a validação
        formula = f'=INDIRECT("\'Dados Consolidados\'!SecaoCompleta" & Y{i})'
        
        dv = DataValidation(type="list", formula1=formula, showDropDown=False)
        dv.error = "Por favor, selecione um valor da lista."
        dv.errorTitle = "Valor Inválido"
        dv.showErrorMessage = True
        
        aba_planilha.add_data_validation(dv)
        dv.add(aba_planilha[f"B{i}"])  # Aplica a validação na célula B{i}

    if "Cadastro de Seção" in wb.sheetnames:
        aba_secao = wb["Cadastro de Seção"]
        adicionar_validacao(aba_secao, "B7:B200", f"'Dados Consolidados'!$AR$2:$AR${len(dados['secao_completa'])}")

    if "Cadastro de Espécie" in wb.sheetnames:
        aba_especie = wb["Cadastro de Espécie"]
        adicionar_validacao(aba_especie, "B7:B200", f"'Dados Consolidados'!$A$2:$A${len(dados['especie_completa'])}")

    """if "Cadastro de Marcas" in wb.sheetnames:
        aba_marcas = wb["Cadastro de Marcas"]
        adicionar_validacao(aba_marcas, "A7:A200", "ÉERRO(CORRESP($A7;'Dados Consolidados'!$AT$1:$AT$200;0))")

    if "Cadastro de Segmento" in wb.sheetnames:
        aba_segmento = wb["Cadastro de Segmento"]
        
        # Criar validação personalizada sem restringir a lista suspensa
        formula = "=ISNA(CORRESP(A7; 'Dados Consolidados'!$AR$1:$AR$200; 0))"
        
        dv = DataValidation(type="custom", formula1=formula)
        dv.error = "O valor digitado não existe na base de segmentos."
        dv.errorTitle = "Valor Inválido"
        dv.showErrorMessage = True

        # Aplicar a validação a todas as células da coluna A (de A7 até A200)
        aba_segmento.add_data_validation(dv)
        for linha in aba_segmento["A7:A200"]:
            for celula in linha:
                dv.add(celula)"""


    max_linhas = max(len(lista) for lista in dados.values()) + 10  

    colunas_com_validacao = ["A", "B", "E", "H", "J", "K", "P"]

    print("Adicionando validação de dados gerais...\n")    
    for chave, coluna in tqdm(mapeamento_colunas.items(), desc="Validação"):
        if coluna in colunas_com_validacao and chave in dados:  
            adicionar_validacao(aba_planilha, f"{coluna}7:{coluna}{max_linhas}", f"'{nome_aba_dados}'!${coluna}$1:${coluna}${len(dados[chave])}")

    wb.save(caminho_arquivo_novo)
    
    tempo_total = time.time() - inicio

    if tempo_total > 60:
        minutos = int(tempo_total // 60)
        segundos = tempo_total % 60
        print(f"Tempo total para preencher planilha: {minutos} minutos e {segundos:.0f} segundos\n")
    else:
        print(f"Tempo total para preencher planilha: {tempo_total:.0f} segundos\n")

caminho_arquivo = 'Cadastros Auto Nextt limpa.xlsx'

if not os.path.exists(caminho_arquivo):
    print(f"Arquivo não encontrado: {caminho_arquivo}")
    exit()

print("Dados preenchidos com sucesso e a planilha original foi excluída.")
