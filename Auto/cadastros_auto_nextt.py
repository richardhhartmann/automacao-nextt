import pandas as pd
import pyodbc
import tkinter as tk
import os
import json
import sys
import openpyxl
import time
import re
import warnings
from tkinter import filedialog
from datetime import datetime
from openpyxl.utils import get_column_letter, column_index_from_string
from contextlib import contextmanager

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

BATCH_SIZE = 100
LINHA_CABECALHO = 3
COLUNA_INICIAL_ADICIONAIS = 'BD'
COLUNA_INICIAL_ADICIONAIS_PEDIDO = 'AN'
DEBUG = True
CODIGOS = 10

def debug_log(message):
    if DEBUG:
        print(f"[DEBUG] {message}")

@contextmanager
def get_db_connection(file_name='conexao_temp.txt'):
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(file_name)
        
        with open(file_path, 'r') as f:
            config = json.load(f)

        driver = config.get('driver', None)
        server = config.get('server', None)
        database = config.get('database', None)
        username = config.get('username', None)
        password = config.get('password', None)
        trusted_connection = config.get('trusted_connection', None)

        if trusted_connection and trusted_connection.lower() == 'yes':
            conn_str = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};Trusted_Connection={trusted_connection}"
        else:
            conn_str = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={username};PWD={password};"

        global connection, cursor
        connection = pyodbc.connect(conn_str)
        connection.autocommit = False
        cursor = connection.cursor()
        debug_log("Conexão com o banco estabelecida")
        
        yield connection, cursor
        
    except Exception as e:
        debug_log(f"ERRO na conexão: {e}")
        sys.exit(1)
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

def get_colunas_adicionais(ws, linha_cabecalho=LINHA_CABECALHO, coluna_inicial=COLUNA_INICIAL_ADICIONAIS):
    colunas = []
    current_col_idx = column_index_from_string(coluna_inicial)
    limite_col_idx = column_index_from_string('CF') 
    
    while current_col_idx <= limite_col_idx:
        col_letter = get_column_letter(current_col_idx)
        valor = ws[f'{col_letter}{linha_cabecalho}'].value
        if pd.isna(valor) or not valor:
            break
        colunas.append(col_letter)
        current_col_idx += 1

    return colunas

def get_colunas_adicionais_pedido(ws, linha_cabecalho=LINHA_CABECALHO, coluna_inicial=COLUNA_INICIAL_ADICIONAIS_PEDIDO):
    colunas = []
    current_col_idx = column_index_from_string(coluna_inicial)

    while True:
        col_letter = get_column_letter(current_col_idx)
        valor = ws[f'{col_letter}{linha_cabecalho}'].value
        if pd.isna(valor) or not valor:
            break
        colunas.append(col_letter)
        current_col_idx += 1

    return colunas

def trata_valor(valor, tipo=int):
    if pd.isna(valor) or str(valor).strip().upper() in ['#N/D', '#N/A', 'N/D', '']:
        return None
    try:
        return tipo(valor)
    except (ValueError, TypeError):
        return None

def processa_produto(ws, linha_excel, df, x):
    if df.iloc[x, 199] != "OK":
        return None

    secao = trata_valor(df.iloc[x, 84]) #
    especie = trata_valor(df.iloc[x, 85]) #
    descricao = str(df.iloc[x, 2])[:50] if pd.notna(df.iloc[x, 2]) else None
    descricao_reduzida = str(df.iloc[x, 3])[:50] if pd.notna(df.iloc[x, 3]) else None
    marca = trata_valor(df.iloc[x, 86]) #
    comprador = trata_valor(df.iloc[x, 87]) #
    und_codigo = trata_valor(df.iloc[x, 88]) #
    classificacao = trata_valor(df.iloc[x, 89]) #
    origem = trata_valor(df.iloc[x, 90]) #
    etiqueta = trata_valor(df.iloc[x, 91]) #

    referencia = (str(int(df.iloc[x, 5])) if isinstance(df.iloc[x, 5], float) and df.iloc[x, 5].is_integer() else str(df.iloc[x, 5])) if pd.notna(df.iloc[x, 5]) else None
    cod_original = str(df.iloc[x, 6]) if pd.notna(df.iloc[x, 6]) else ''
    
    ativo = 1
    venda = float(str(df.iloc[x, 12]).replace(',', '.')) if pd.notna(df.iloc[x, 12]) else None
    icms = float(df.iloc[x, 13]) if pd.notna(df.iloc[x, 13]) else None
    ipi = float(df.iloc[x, 14]) if pd.notna(df.iloc[x, 14]) else None
    ipr_codigo_barra = trata_valor(df.iloc[x, 16])
    
    data = datetime.now()
    
    # Atributos adicionais (mantido da versão anterior)
    atributos_adicionais = []
    colunas_adicionais = get_colunas_adicionais(ws)
    
    for idx, col in enumerate(colunas_adicionais, start=3):
        valor = ws[f'{col}{linha_excel}'].value
        if pd.notna(valor):
            atributos_adicionais.append((secao, especie, None, idx, str(valor), None))

    # Nova lógica para cores e tamanhos
    cores = []
    tamanhos = []
    
    # Coleta de cores (colunas R a Z)
    for col_letra in ['R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 
                     'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ']:
        valor = ws[f'{col_letra}{linha_excel}'].value
        if pd.notna(valor) and str(valor).strip() != '':
            cores.append(str(valor).strip())
    
    # Coleta de tamanhos (colunas AK a BC)
    for col_letra in ['AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 
                     'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC']:
        valor = ws[f'{col_letra}{linha_excel}'].value
        if pd.notna(valor) and str(valor).strip() != '':
            tamanhos.append(str(valor).strip())
    
    variacoes = []
    ipr_codigo = 1
    
    if cores and tamanhos:
        for cor in cores:
            for tamanho in tamanhos:
                variacoes.append({
                    'ipr_codigo': ipr_codigo,
                    'ipr_codigo_barra': ipr_codigo_barra,
                    'cor': cor,
                    'tamanho': tamanho
                })
                ipr_codigo += 1
    elif cores:
        for cor in cores:
            variacoes.append({
                'ipr_codigo': ipr_codigo,
                'ipr_codigo_barra': ipr_codigo_barra,
                'cor': cor,
                'tamanho': None
            })
            ipr_codigo += 1
    elif tamanhos:
        for tamanho in tamanhos:
            variacoes.append({
                'ipr_codigo': ipr_codigo,
                'ipr_codigo_barra': ipr_codigo_barra,
                'cor': None,
                'tamanho': tamanho
            })
            ipr_codigo += 1
    
    return {
        'secao': secao,
        'especie': especie,
        'descricao': descricao,
        'descricao_reduzida': descricao_reduzida,
        'marca': marca,
        'comprador': comprador,
        'und_codigo': und_codigo,
        'classificacao': classificacao,
        'origem': origem,
        'etiqueta': etiqueta,
        'referencia': referencia,
        'cod_original': cod_original,
        'ativo': ativo,
        'venda': venda,
        'icms': icms,
        'ipi': ipi,
        'data': data,
        'atributos_adicionais': atributos_adicionais,
        'variacoes': variacoes,
        'linha_excel': linha_excel
    }

def processa_pedido(ws, linha_excel, df, x):
    try:
        col_map = {
            'cod_original': 'cod_original',
            'fornecedor': 'fornecedor',
            'comprador': 'comprador',
            'dt_entrega_inicial': 'dt_entrega_inicial',
            'dt_entrega_final': 'dt_entrega_final',
            'condicao_pagamento': 'condicao_pagamento',
            'qualidade': 'qualidade',
            'parcelas': 'parcelas',
            'observacao': 'observacao',
            'atributo': 'atributo'
        }
        
        data = {}
        has_value = False
        
        for field, col in col_map.items():
            try:
                val = df.iloc[x][col]
                if pd.notna(val) and (not isinstance(val, (int, float)) or (isinstance(val, (int, float)) and val != 0)):
                    if field in ['dt_entrega_inicial', 'dt_entrega_final', 'observacao']:
                        data[field] = str(val)[:50]
                    else:
                        data[field] = trata_valor(val)
                    has_value = True
                else:
                    data[field] = None
            except:
                data[field] = None
        
        data['formas_pagamento'] = None
        data['ped_status'] = 'D'
        
        return data if has_value else None
        
    except Exception as e:
        return None
    
def verificar_duplicados(cursor, referencias_marcas):
    if not referencias_marcas:
        return set()
    
    cursor.execute("CREATE TABLE #TempDuplicados (Referencia NVARCHAR(255), Marca INT)")
    cursor.executemany("INSERT INTO #TempDuplicados (Referencia, Marca) VALUES (?, ?)", referencias_marcas)
    
    cursor.execute("""
        SELECT t.Referencia, t.Marca 
        FROM #TempDuplicados t
        INNER JOIN tb_produto p ON p.prd_referencia_fornec = t.Referencia AND p.mar_codigo = t.Marca
    """)
    
    duplicados = set((row.Referencia, row.Marca) for row in cursor.fetchall())
    cursor.execute("DROP TABLE #TempDuplicados")
    
    debug_log(f"Encontrados {len(duplicados)} produtos duplicados")
    return duplicados

def cadastrar_produto(excel):
    start_total = time.time()

    if not excel:
        return
    
    try:
        debug_log("Carregando arquivo Excel")
        wb = openpyxl.load_workbook(excel, data_only=True)
            
        ws = wb["Cadastro de Produtos"]
        
        df = pd.read_excel(excel, sheet_name="Cadastro de Produtos", skiprows=6, header=None)

        df = df.dropna(how='all')
        
        df.columns = [
            "secao", "especie", "descricao", "descricao_reduzida", "marca", "referencia",
            "cod_original", "comprador", "ativo", "unidade", "classificacao", "origem",
            "venda", "icms", "ipi", "etiqueta", "coluna17", "coluna18", "coluna19", "coluna20",
            "coluna21", "coluna22", "coluna23", "coluna24", "coluna25", "coluna26", "coluna27",
            "coluna28", "coluna29", "coluna30", "coluna31"
        ] + [f"coluna{i}" for i in range(32, len(df.columns) + 1)]

        colunas_numericas = ['secao', 'especie', 'cod_original', 'classificacao', 'venda', 'origem', 
                            'icms', 'ipi', 'etiqueta', 'comprador', 'coluna25', 'coluna26', 'coluna27',
                            'coluna28', 'coluna29', 'coluna30', 'coluna31', 'coluna54', 'coluna55',
                            'coluna56', 'coluna57', 'coluna58', 'coluna59', 'coluna60', 'coluna61',
                            'coluna62']
        
        for col in colunas_numericas:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype('int32')

        produtos = []
        for x in range(len(df)):
            linha_excel = x + 7
            produto = processa_produto(ws, linha_excel, df, x)
            if produto:
                produtos.append(produto)
        
        debug_log(f"Total de produtos a processar: {len(produtos)}")

        with get_db_connection('conexao_temp.txt') as (connection, cursor):
            duplicata = 0
            produtos_inseridos = 0
            variacoes_inseridas = 0
            
            referencias_marcas = [(p['referencia'], p['marca']) for p in produtos if p['referencia'] and p['marca']]
            duplicados = verificar_duplicados(cursor, referencias_marcas)
            
            for i, produto in enumerate(produtos, 1):
                if (produto['referencia'], produto['marca']) in duplicados:
                    duplicata += 1
                    continue
                
                cursor.execute("SELECT MAX(prd_codigo) FROM tb_produto WHERE sec_codigo = ? AND esp_codigo = ?", produto['secao'], produto['especie'])
                resultado = cursor.fetchone()
                prd_codigo = (resultado[0] or 0) + 1
                
                parametros_produto = (
                    produto['secao'], produto['especie'], prd_codigo, produto['descricao'], 
                    produto['descricao_reduzida'], produto['marca'], produto['data'], None, None, None,
                    0, 0, 0, produto['cod_original'], produto['ativo'], 0, None, produto['referencia'],
                    None, produto['classificacao'], produto['comprador'], produto['und_codigo'], 
                    produto['venda'], produto['origem'], produto['icms'], produto['ipi'], 
                    produto['etiqueta'], None, None, None, 1, None, None, None, None, None
                )
                
                print("\nDados processados:")
                pd.set_option('display.max_columns', None)
                pd.set_option('display.max_colwidth', None)
                print(df)

                cursor.execute("""
                    INSERT INTO tb_produto
                    (sec_codigo, esp_codigo, prd_codigo, prd_descricao, prd_descricao_reduzida,
                    mar_codigo, prd_data_cadastro, prd_unidade, prd_data_ultima_compra,
                    prd_data_ultima_entrega, prd_custo_medio, prd_preco_medio, prd_aliquota_icms,
                    prd_codigo_original, prd_ativo, prd_ultimo_custo, prd_arquivo_foto,
                    prd_referencia_fornec, prd_tipo_tributacao, clf_codigo, usu_codigo_comprador,
                    und_codigo, prd_valor_venda, prd_origem, prd_percentual_icms, prd_percentual_ipi,
                    etq_codigo_padrao, usu_codigo_cadastro, sec_codigo_r, esp_codigo_r,
                    prd_permite_comprar, prd_valor_unidade_conversao, udc_codigo, und_codigo_conversao,
                    prd_iat, prd_ippt)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
                            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, parametros_produto)
                
                for atributo in produto['atributos_adicionais']:
                    sec, esp, _, tpa, desc, _ = atributo

                    cursor.execute("""
                        SELECT COUNT(*) 
                        FROM INFORMATION_SCHEMA.COLUMNS 
                        WHERE TABLE_NAME = 'tb_atributo_produto' 
                        AND COLUMN_NAME = 'prr_codigo'
                    """)
                    has_prr_codigo = cursor.fetchone()[0] > 0
                    
                    if has_prr_codigo:
                        try:
                            cursor.execute("""
                                INSERT INTO tb_atributo_produto 
                                (sec_codigo, esp_codigo, prd_codigo, tpa_codigo, apr_descricao, prr_codigo)
                                VALUES (?, ?, ?, ?, ?, NULL)
                            """, sec, esp, prd_codigo, tpa, desc)
                        except pyodbc.IntegrityError:
                            pass
                    else:
                        cursor.execute("""
                            INSERT INTO tb_atributo_produto 
                            (sec_codigo, esp_codigo, prd_codigo, tpa_codigo, apr_descricao)
                            VALUES (?, ?, ?, ?, ?)
                        """, (sec, esp, prd_codigo, tpa, desc))
                
                cursor.execute("""SELECT COUNT(*) 
                    FROM INFORMATION_SCHEMA.COLUMNS 
                    WHERE TABLE_NAME = 'tb_atributo_produto' 
                    AND COLUMN_NAME = 'ipr_gtin'""")
                
                has_ipr_gtin = cursor.fetchone()[0] > 0

                for variacao in produto['variacoes']:
                    if has_ipr_gtin:
                        cursor.execute("""
                            INSERT INTO tb_item_produto 
                            (sec_codigo, esp_codigo, prd_codigo, ipr_codigo, ipr_codigo_barra, ipr_preco_promocional, ipr_gtin)
                            VALUES (?, ?, ?, ?, ?, 0, NULL)
                        """, produto['secao'], produto['especie'], prd_codigo, variacao['ipr_codigo'], variacao['ipr_codigo_barra'])
                    else:
                        cursor.execute("""
                            INSERT INTO tb_item_produto 
                            (sec_codigo, esp_codigo, prd_codigo, ipr_codigo, ipr_codigo_barra, ipr_preco_promocional)
                            VALUES (?, ?, ?, ?, ?, 0)
                        """, produto['secao'], produto['especie'], prd_codigo, variacao['ipr_codigo'], variacao['ipr_codigo_barra'])

                    if variacao['cor']:
                        cursor.execute("""
                            INSERT INTO tb_atributo_item_produto 
                            (sec_codigo, esp_codigo, prd_codigo, ipr_codigo,
                            tpa_codigo, aip_descricao, aip_ordem, aip_descricao_fornec)
                            VALUES (?, ?, ?, ?, 1, ?, ?, NULL)
                        """, produto['secao'], produto['especie'], prd_codigo, variacao['ipr_codigo'], 
                            variacao['cor'], variacao['ipr_codigo'])
                    
                    if variacao['tamanho']:
                        cursor.execute("""
                            INSERT INTO tb_atributo_item_produto 
                            (sec_codigo, esp_codigo, prd_codigo, ipr_codigo,
                            tpa_codigo, aip_descricao, aip_ordem, aip_descricao_fornec)
                            VALUES (?, ?, ?, ?, 2, ?, ?, NULL)
                        """, produto['secao'], produto['especie'], prd_codigo, variacao['ipr_codigo'], 
                            variacao['tamanho'], variacao['ipr_codigo'])
                    
                    variacoes_inseridas += 1
                
                produtos_inseridos += 1
                
                if i % BATCH_SIZE == 0:
                    connection.commit()
                    debug_log(f"Lote {i//BATCH_SIZE} commitado ({i} produtos processados)")
            
            connection.commit()
            
            tempo_total = time.time() - start_total
            debug_log(f"Processo concluído em {tempo_total:.2f} segundos")
            print("\nResumo:")
            print(f"- Total de produtos processados: {len(produtos)}")
            print(f"- Produtos inseridos: {produtos_inseridos}")
            print(f"- Variações inseridas: {variacoes_inseridas}")
            print(f"- Produtos duplicados (ignorados): {duplicata}")
            print(f"- Tempo médio por produto: {tempo_total/max(1, produtos_inseridos):.2f}s")
            
    except Exception as e:
        debug_log(f"ERRO: {str(e)}")
        if 'connection' in locals():
            connection.rollback()
    finally:
        if 'wb' in locals():
            wb.close()

def cadastrar_pedido(excel):
    debug_log("Iniciando processo de cadastro de pedido")
    start_total = time.time()
    
    if not excel:
        return []

    try:
        debug_log("Carregando arquivo Excel")
        df = pd.read_excel(excel, sheet_name="Cadastro de Pedidos", skiprows=6, header=None)
        df = df.dropna(how='all')
        
        if df.empty:
            debug_log("Nenhum dado válido encontrado")
            return []

        num_colunas = df.shape[1]

        nomes_colunas = [
            "cod_original", "fornecedor", "comprador", "dt_entrega_inicial", "dt_entrega_final",
            "condicao_pagamento", "qualidade", "parcelas", "observacao", "formas_pagamento",
            "atributo"
        ]
        
        nomes_colunas += [f"cod{i}" for i in range(10)]
        nomes_colunas += [f"pag{i}" for i in range(10)]
        
        coluna_atual = 32
        while len(nomes_colunas) < num_colunas:
            if coluna_atual != 703:
                nomes_colunas.append(f"coluna{coluna_atual}")
            coluna_atual += 1

        if len(nomes_colunas) > num_colunas:
            nomes_colunas = nomes_colunas[:num_colunas]
        elif len(nomes_colunas) < num_colunas:
            nomes_colunas += [f"extra_{i}" for i in range(num_colunas - len(nomes_colunas))]
        
        df.columns = nomes_colunas

        df = df[df['fornecedor'].notna() & (df['fornecedor'] != 0)]
        
        colunas_numericas = ['cod_original'] + \
                           [f"coluna{i}" for i in [705, 706, 707, 708] if f"coluna{i}" in df.columns] + \
                           [f"cod{i}" for i in range(10)] + \
                           [f"pag{i}" for i in range(10)]
        
        for col in colunas_numericas:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype('int32')

        for col_data in ['dt_entrega_inicial', 'dt_entrega_final']:
            if col_data in df.columns:
                df[col_data] = pd.to_datetime(df[col_data], errors='coerce')
                df[col_data] = df[col_data].dt.strftime('%Y-%m-%d %H:%M:%S')

        pedidos = []
        for _, row in df.iterrows():
            pedido = {
                'cod_original': row['cod_original'],
                'fornecedor': str(row['fornecedor'])[:50] if pd.notna(row['fornecedor']) else None,
                'comprador': row.get('comprador', ''),
                'dt_entrega_inicial': row.get('dt_entrega_inicial'),
                'dt_entrega_final': row.get('dt_entrega_final'),
                'condicao_pagamento': row.get('condicao_pagamento', 0),
                'qualidade': row.get('qualidade', ''),
                'parcelas': row.get('parcelas', 0),
                'observacao': str(row['observacao'])[:50] if pd.notna(row['observacao']) else None,
                'atributo': row.get('atributo', 0),
                'formas_pagamento': None,
                'ped_status': 'D'
            }
            pedidos.append(pedido)

        debug_log(f"Total de pedidos válidos processados: {len(pedidos)}")
        
        """print("\nDados processados:")
        pd.set_option('display.max_columns', None)
        pd.set_option('display.max_colwidth', None)
        print(df)"""
        
        with get_db_connection('conexao_temp.txt') as (conn, cursor):
            try:
                cursor.execute("SELECT ISNULL(MAX(ped_codigo), 0) + 1 FROM tb_pedido")
                proximo_codigo = cursor.fetchone()[0]

                pedidos_inseridos = 0
                for _, row in df.iterrows():
                    if pd.notna(row['fornecedor']) and row['fornecedor'] != 0:
                        try:
                            def format_sql_date(date_str):
                                try:
                                    dt = pd.to_datetime(date_str, errors='coerce')
                                    return dt.strftime('%Y-%m-%d %H:%M:%S') if not pd.isna(dt) else None
                                except:
                                    return None
                            data_atual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                            fornecedor = str(df.iloc[_, 702]).strip()
                            comprador = str(df.iloc[_, 703])[:50] if pd.notna(df.iloc[_, 703]) else None

                            data_entrega_inicial = format_sql_date(row['dt_entrega_inicial'])
                            data_entrega_final = format_sql_date(row['dt_entrega_final'])

                            try:
                                cpg_codigo = int(float(row.get('condicao_pagamento', 0)))
                            except (ValueError, TypeError):
                                cpg_codigo = None

                            cursor.execute("""SELECT COUNT(*) 
                    FROM INFORMATION_SCHEMA.COLUMNS 
                    WHERE TABLE_NAME = 'tb_pedido' 
                    AND COLUMN_NAME = 'ped_permite_divergencia_entrada_nf'
                                           """)
                            
                            has_ped_permite_divergencia_entrada_nf = cursor.fetchone()[0] > 0
                            
                            if has_ped_permite_divergencia_entrada_nf:
                                sql = """
                                INSERT INTO tb_pedido(
                                    ped_codigo, pes_codigo, usu_codigo_comprador, 
                                    ped_data_emissao, ped_data_entrega_inicial, ped_data_entrega_final, 
                                    ped_status, ped_observacao, ped_qtde_total, 
                                    ped_valor_total, ped_qtde_entregue_total, ped_custo_medio, 
                                    ped_codigo_original, ped_qualidade, ped_comissao_compras, 
                                    cpg_codigo, ped_data_cadastro, usu_codigo_cadastro, 
                                    ped_data_ult_alteracao, ped_permite_divergencia_entrada_nf
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """
                                
                                params = (
                                    int(proximo_codigo), 
                                    fornecedor,
                                    comprador,
                                    data_atual,
                                    data_entrega_inicial if data_entrega_inicial else None,
                                    data_entrega_final if data_entrega_final else None,
                                    'D',  # Status Digitado
                                    str(row['observacao']).strip()[:50] if pd.notna(row['observacao']) else None,
                                    0,  # ped_qtde_total
                                    0,  # ped_valor_total
                                    0,  # ped_qtde_entregue_total
                                    0,  # ped_custo_medio
                                    int(row['cod_original']),
                                    str(row.get('qualidade', '')).strip(),
                                    0,  # ped_comissao_compras
                                    cpg_codigo,
                                    data_atual,
                                    None,
                                    data_atual,
                                    0  # ped_permite_divergencia_entrada_nf
                                )
                            else:
                                sql = """
                                INSERT INTO tb_pedido(
                                    ped_codigo, pes_codigo, usu_codigo_comprador, 
                                    ped_data_emissao, ped_data_entrega_inicial, ped_data_entrega_final, 
                                    ped_status, ped_observacao, ped_qtde_total, 
                                    ped_valor_total, ped_qtde_entregue_total, ped_custo_medio, 
                                    ped_codigo_original, ped_qualidade, ped_comissao_compras, 
                                    cpg_codigo, ped_data_cadastro, usu_codigo_cadastro, 
                                    ped_data_ult_alteracao
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """
                                
                                params = (
                                    int(proximo_codigo), 
                                    fornecedor,
                                    comprador,
                                    data_atual,
                                    data_entrega_inicial if data_entrega_inicial else None,
                                    data_entrega_final if data_entrega_final else None,
                                    'D',  # Status Digitado
                                    str(row['observacao']).strip()[:50] if pd.notna(row['observacao']) else None,
                                    0,  # ped_qtde_total
                                    0,  # ped_valor_total
                                    0,  # ped_qtde_entregue_total
                                    0,  # ped_custo_medio
                                    int(row['cod_original']),
                                    str(row.get('qualidade', '')).strip(),
                                    0,  # ped_comissao_compras
                                    cpg_codigo,
                                    data_atual,
                                    comprador,  # usu_codigo_cadastro
                                    data_atual
                                )

                            debug_log(f"Parâmetros lidos: {params}")

                            cursor.execute(sql, params)
                            conn.commit()
                            proximo_codigo += 1
                            pedidos_inseridos += 1
                            pedido_atual = proximo_codigo-1
                            debug_log(f"Pedido {pedido_atual} cadastrado com sucesso!")

                            formas_de_pagamento = []
                            coluna_inicial_pagamentos = 708 - 1  # Convertendo para 0-based index do pandas

                            if df.shape[1] > coluna_inicial_pagamentos:
                                for col_idx in range(coluna_inicial_pagamentos, df.shape[1]):
                                    val = row.iloc[col_idx]
                                    if pd.notna(val) and val != 0:
                                        try:
                                            formas_de_pagamento.append(int(val))
                                        except (ValueError, TypeError):
                                            continue

                            formas_de_pagamento = ','.join(map(str, formas_de_pagamento)) if formas_de_pagamento else None

                            if formas_de_pagamento:
                                try:
                                    cursor.execute("DELETE FROM tb_pedido_tipo_documento WHERE ped_codigo = ?", (pedido_atual,))
                                    
                                    for tid_codigo in formas_de_pagamento.split(','):
                                        cursor.execute("""
                                            INSERT INTO tb_pedido_tipo_documento(ped_codigo, tid_codigo) 
                                            VALUES (?, ?)
                                            """, (pedido_atual, int(tid_codigo)))
                                    
                                    conn.commit()
                                    debug_log(f"Formas de pagamento {formas_de_pagamento} inseridas para pedido {pedido_atual}")
                                except Exception as e:
                                    conn.rollback()
                                    debug_log(f"Erro ao inserir formas de pagamento para pedido {pedido_atual}: {str(e)}")

                            with get_db_connection('conexao_temp.txt') as (conn, cursor):
                                cursor.execute("SELECT fil_codigo FROM tb_filial ORDER BY fil_codigo")
                                filiais = [row[0] for row in cursor.fetchall()]
                                debug_log(f"Encontradas {len(filiais)} filiais no sistema")

                            for i in range(10):
                                cod_col = f"cod{i}"
                                pag_col = f"pag{i}"
                                
                                if cod_col in row and pd.notna(row[cod_col]) and row[cod_col] != 0:
                                    try:
                                        codigo = str(int(row[cod_col])).zfill(9)
                                        sec_codigo = int(codigo[:3])
                                        esp_codigo = int(codigo[3:5])
                                        prd_codigo = int(codigo[5:9])
                                        
                                        valor_custo = float(row[pag_col]) if pag_col in row and pd.notna(row[pag_col]) else 0
                                        
                                        coluna_inicial_filial = 39 + (i * len(filiais))
                                        
                                        for filial_idx, fil_codigo in enumerate(filiais):
                                            coluna_fator = coluna_inicial_filial + filial_idx
                                            
                                            fator_filial = 0  
                                            if coluna_fator < len(row):
                                                try:
                                                    fator_filial = int(row.iloc[coluna_fator]) if pd.notna(row.iloc[coluna_fator]) else 0
                                                except (ValueError, TypeError):
                                                    fator_filial = 0
                                            
                                            if filial_idx == 0:
                                                cursor.execute("""
                                                    INSERT INTO tb_pedido_produto(
                                                        ped_codigo, sec_codigo, esp_codigo, prd_codigo, 
                                                        ppr_qtde_pedido, ppr_qtde_entregue, ppr_custo_medio, 
                                                        ppr_valor_desconto, ppr_percentual_icms, ppr_percentual_ipi, 
                                                        ppr_valor_icms, ppr_valor_ipi, ppr_sequencial, 
                                                        ppr_multiplicador_unidade, ppr_data_entrega_inicial, 
                                                        ppr_data_entrega_final, ppr_valor_icms_st
                                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, 0, 0, 0, 0, 0, NULL, 0, 
                                                            '2025-04-16 13:21:00.000', '2025-04-16 13:21:00.000', NULL)
                                                    """, 
                                                    (pedido_atual, sec_codigo, esp_codigo, prd_codigo,
                                                    1, 10, valor_custo)
                                                )
                                            
                                            cursor.execute("""
                                                INSERT INTO tb_item_pedido(
                                                    ped_codigo, sec_codigo, esp_codigo, prd_codigo, 
                                                    ipr_codigo, fil_codigo, ipd_pack, ipd_qtde_pedido, 
                                                    ipd_qtde_entregue, ipd_valor_custo, ipd_valor_desconto, 
                                                    ipd_fator_grade, ipd_fator_filial, ipd_valor_icms, 
                                                    ipd_valor_ipi
                                                ) VALUES (?, ?, ?, ?, 1, ?, 1, ?, 0, ?, 0, 1, ?, 0, 0)
                                                """,
                                                (pedido_atual, sec_codigo, esp_codigo, prd_codigo,
                                                fil_codigo,
                                                fator_filial,  # esse valor provavelmente é dado pelo resultado da divisão de fator_filial por ipd_qtde_pedido
                                                valor_custo,
                                                fator_filial)
                                            )
                                            
                                            debug_log(f"Produto {i} - Filial {fil_codigo}: Coluna {coluna_fator} = Fator {fator_filial}")
                                        
                                        conn.commit()
                                        debug_log(f"Produto {codigo} processado com {len(filiais)} fatores filiais")
                                        
                                    except Exception as e:
                                        conn.rollback()
                                        debug_log(f"Erro ao processar produto {i}: {str(e)}")
                                        debug_log(f"Valores problemáticos - Código: {row[cod_col]}, Custo: {valor_custo}")
                                        
                        except Exception as e:
                            conn.rollback()
                            debug_log(f"Erro ao inserir pedido {pedido_atual}: {str(e)}")
                            debug_log(f"Tipo dos parâmetros: {tuple(type(p) for p in params)}")
                            debug_log(f"Valores problemáticos: {params}")

                debug_log(f"Total de pedidos cadastrados: {pedidos_inseridos}")
                return pedidos_inseridos

            except Exception as e:
                conn.rollback()
                debug_log(f"Erro durante o processamento no banco: {str(e)}")
                return 0

    except Exception as e:
        debug_log(f"Erro durante o processamento: {str(e)}")
        return 0
    finally:
        debug_log(f"Tempo total: {time.time() - start_total:.2f}s")