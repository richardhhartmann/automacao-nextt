import openpyxl
import pyodbc
import json
from tkinter import Tk, filedialog

def selecionar_arquivo_excel():
    root = Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Selecione a planilha Excel",
        filetypes=[("Arquivos do Excel", "*.xlsx *.xlsm")]
    )

def conectar_banco():
    with open("conexao_temp.txt", "r", encoding="utf-8") as f:
        config = json.load(f)

    parts = [
        f"DRIVER={{{config['driver']}}}",
        f"SERVER={config['server']}",
        f"DATABASE={config['database']}"
    ]

    if config.get("trusted_connection", "no").lower() == "yes":
        parts.append("Trusted_Connection=yes")
    else:
        parts.append(f"UID={config['username']}")
        parts.append(f"PWD={config['password']}")

    conn_str = ";".join(parts)
    return pyodbc.connect(conn_str)

def importacao_dados(caminho):
    if not caminho:
        print("Nenhum arquivo selecionado.")
        return

    wb = openpyxl.load_workbook(caminho, data_only=True)
    
    try:
        conn = conectar_banco()
        cursor = conn.cursor()

        # MARCAS
        if "Cadastro de Marcas" in wb.sheetnames:
            ws_marcas = wb["Cadastro de Marcas"]
            cursor.execute("SELECT ISNULL(MAX(mar_codigo), 0) + 1 FROM tb_marca")
            mar_codigo = cursor.fetchone()[0] or 1

            sql_insert_marcas = "INSERT INTO tb_marca (mar_codigo, mar_descricao) VALUES (?, ?)"
            dados_marca = []

            for row in ws_marcas.iter_rows(min_row=7, max_row=1007, min_col=1, max_col=1):
                mar_descricao = row[0].value
                if mar_descricao:
                    dados_marca.append((mar_codigo, mar_descricao))
                    mar_codigo += 1

            if dados_marca:
                cursor.executemany(sql_insert_marcas, dados_marca)
                conn.commit()
                print(f"{len(dados_marca)} marcas inseridas com sucesso!")
            else:
                print("A planilha 'Cadastro de Marcas' não contém dados válidos para importação.")
        else:
            print("A aba 'Cadastro de Marcas' não foi encontrada.")

        # SEGMENTOS
        if "Cadastro de Segmento" in wb.sheetnames:
            ws_segmento = wb["Cadastro de Segmento"]
            cursor.execute("SELECT ISNULL(MAX(seg_codigo), 0) + 1 FROM tb_segmento")
            seg_codigo = cursor.fetchone()[0] or 1

            sql_insert_segmento = "INSERT INTO tb_segmento (seg_codigo, seg_descricao, ram_codigo) VALUES (?, ?, ?)"
            dados_segmento = []

            for row in ws_segmento.iter_rows(min_row=7, max_row=1007, min_col=1, max_col=1):
                seg_descricao = row[0].value
                if seg_descricao:
                    dados_segmento.append((seg_codigo, seg_descricao, 1))
                    seg_codigo += 1

            if dados_segmento:
                cursor.executemany(sql_insert_segmento, dados_segmento)
                conn.commit()
                print(f"{len(dados_segmento)} segmentos inseridos com sucesso!")
            else:
                print("A planilha 'Cadastro de Segmento' não contém dados válidos para importação.")
        else:
            print("A aba 'Cadastro de Segmento' não foi encontrada.")

        # SEÇÕES
        if "Cadastro de Secao" in wb.sheetnames:
            ws_secao = wb["Cadastro de Secao"]
            cursor.execute("SELECT ISNULL(MAX(sec_codigo), 0) + 1 FROM tb_secao")
            sec_codigo = cursor.fetchone()[0] or 1

            sql_insert_secao = """
                INSERT INTO tb_secao (
                    sec_codigo, sec_descricao, seg_codigo, 
                    sec_permite_item_produto, sec_ativo, usu_codigo_comprador, clf_codigo
                ) VALUES (?, ?, ?, 1, 1, 1, NULL)
            """
            dados_secao = []

            for row in ws_secao.iter_rows(min_row=7, max_row=1007, min_col=1, max_col=3):
                sec_descricao = row[0].value
                seg_codigo_ref = row[2].value
                if sec_descricao and seg_codigo_ref:
                    dados_secao.append((sec_codigo, sec_descricao, seg_codigo_ref))
                    sec_codigo += 1

            if dados_secao:
                cursor.executemany(sql_insert_secao, dados_secao)
                conn.commit()
                print(f"{len(dados_secao)} seções inseridas com sucesso!")
            else:
                print("A planilha 'Cadastro de Secao' não contém dados válidos para importação.")
        else:
            print("A aba 'Cadastro de Secao' não foi encontrada.")

        # ESPÉCIES
        if "Cadastro de Especie" in wb.sheetnames:
            ws_especie = wb["Cadastro de Especie"]
            cursor.execute("SELECT ISNULL(MAX(esp_codigo), 0) + 1 FROM tb_especie")
            esp_codigo = cursor.fetchone()[0] or 1

            sql_insert_especie = """
                INSERT INTO tb_especie (
                    sec_codigo, esp_codigo, esp_descricao, 
                    esp_granel, esp_aliquota_icms, esp_ativo, usu_codigo_comprador, clf_codigo
                ) VALUES (?, ?, ?, 0, NULL, 1, NULL, NULL)
            """
            dados_especie = []

            for row in ws_especie.iter_rows(min_row=7, max_row=1007, min_col=1, max_col=3):
                esp_descricao = row[0].value
                sec_codigo_ref = row[2].value
                if esp_descricao and sec_codigo_ref:
                    dados_especie.append((sec_codigo_ref, esp_codigo, esp_descricao))
                    esp_codigo += 1

            if dados_especie:
                cursor.executemany(sql_insert_especie, dados_especie)
                conn.commit()
                print(f"{len(dados_especie)} espécies inseridas com sucesso!")
            else:
                print("A planilha 'Cadastro de Especie' não contém dados válidos para importação.")
        else:
            print("A aba 'Cadastro de Especie' não foi encontrada.")

        cursor.close()
        conn.close()

    except Exception as e:
        print("Erro ao acessar o banco de dados:", e)