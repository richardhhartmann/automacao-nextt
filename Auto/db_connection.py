import os
import pyodbc
import openpyxl
import shutil
import time
import json
import sys
from openpyxl.worksheet.datavalidation import DataValidation
from tqdm import tqdm

def get_connection_from_file(file_name='conexao_temp.txt'):
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(script_dir, '..', file_name)
        
        with open(file_path, 'r') as f:
            config = json.load(f)

        driver = config.get('driver', None)
        server = config.get('server', None)
        database = config.get('database', None)
        username = config.get('username', None)
        password = config.get('password', None)
        trusted_connection = config.get('trusted_connection', None)

        if trusted_connection and trusted_connection.lower() == 'yes':
            string_connection = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};Trusted_Connection={trusted_connection}"
        else:
            string_connection = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={username};PWD={password};"

        return pyodbc.connect(string_connection)

    except Exception as e:
        print(f"Erro ao conectar ao banco de dados: {e}")
        sys.exit(1)

def preencher_planilha(caminho_arquivo, cancelar_evento):
    if cancelar_evento.is_set():
        print("Processo cancelado antes de iniciar.")
        return

    caminho_arquivo_novo = caminho_arquivo.replace("Cadastros Auto Nextt limpa", "Cadastros Auto Nextt")
    shutil.copy(caminho_arquivo, caminho_arquivo_novo)

    inicio = time.time()
    wb = openpyxl.load_workbook(caminho_arquivo_novo)

    if cancelar_evento.is_set():
        print("Processo cancelado após carregar workbook.")
        return

    connection = get_connection_from_file('conexao_temp.txt')
    cursor = connection.cursor()
    cursor.execute("SELECT TOP 1 emp_descricao FROM tb_empresa")
    empresa_nome = cursor.fetchone()[0]
    connection.close()

    if cancelar_evento.is_set():
        print("Processo cancelado após obter nome da empresa.")
        return

    abas = [
        "Cadastro de Produtos",
        "Cadastro de Pedidos",
        "Cadastro de Marcas",
        "Cadastro de Segmento",
        "Cadastro de Secao",
        "Cadastro de Especie"
    ]

    for aba_nome in abas:
        if cancelar_evento.is_set():
            print("Processo cancelado durante atualização das abas.")
            return
        try:
            aba = wb[aba_nome]
            aba['A2'] = f"Cadastro de {aba_nome.split(' ')[2]} {empresa_nome}"
        except KeyError:
            print(f"A aba '{aba_nome}' não foi encontrada.")

    print("Identificando colunas obrigatórias...")

    connection = get_connection_from_file('conexao_temp.txt')
    cursor = connection.cursor()

    cursor.execute("""
        SELECT COLUMN_NAME 
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_NAME = 'tb_produto' 
        AND IS_NULLABLE = 'NO'
    """)
    colunas_obrigatorias = {row.COLUMN_NAME for row in cursor.fetchall()}
    colunas_obrigatorias.update(["und_codigo", "clf_codigo", "prd_origem"])

    cursor.execute("""
        SELECT COLUMN_NAME 
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_NAME = 'tb_atributo_produto' 
        AND IS_NULLABLE = 'NO'
    """)
    colunas_tb_atributo_produto = {row.COLUMN_NAME for row in cursor.fetchall()}

    aba_planilha = wb["Cadastro de Produtos"]
    linha_titulo = 3
    linha_obrigatorio = 4
    ultima_coluna = 17

    if 'apr_descricao' in colunas_tb_atributo_produto:
        aba_planilha.cell(row=linha_obrigatorio, column=26, value="Obrigatorio")

    mapeamento_colunas = {
        "sec_codigo": "Seção",
        "esp_codigo": "Espécie",
        "prd_descricao": "Descrição",
        "prd_descricao_reduzida": "Descrição Reduzida",
        "mar_codigo": "Marca",
        "prd_referencia_fornec": "Referência do Fornecedor",
        "prd_codigo_original": "Código Original",
        "usu_codigo_comprador": "Comprador",
        "und_codigo": "Unidade", 
        "clf_codigo": "Classificação Fiscal",
        "prd_origem": "Origem",
        "prd_valor_venda": "Valor de Venda",
        "prd_percentual_icms": "% ICMS",
        "prd_percentual_ipi": "% IPI",
        "etq_codigo_padrao": "Etiqueta Padrão"
    }

    for col in range(1, ultima_coluna + 1):
        if cancelar_evento.is_set():
            print("Processo cancelado durante marcação de colunas obrigatórias.")
            return

        nome_coluna_excel = aba_planilha.cell(row=linha_titulo, column=col).value
        if nome_coluna_excel is None:
            nome_coluna_excel = aba_planilha.cell(row=linha_titulo, column=col - 1).value  

        for col_sql, col_excel in mapeamento_colunas.items():
            if nome_coluna_excel == col_excel and col_sql in colunas_obrigatorias:
                aba_planilha.cell(row=linha_obrigatorio, column=col, value="Obrigatorio")

    print("Atualizando validação de dados para espécies...")
    for i in range(7, 1008):
        if cancelar_evento.is_set():
            print("Processo cancelado durante validação de dados.")
            return

        formula = f'=INDIRECT("\'Dados Consolidados\'!SecaoCompleta" & BC{i})'
        dv = DataValidation(type="list", formula1=formula, showDropDown=False)
        dv.error = "Por favor, selecione um valor da lista."
        dv.errorTitle = "Valor Inválido"
        dv.showErrorMessage = True

        aba_planilha.add_data_validation(dv)
        dv.add(aba_planilha[f"B{i}"])

    if cancelar_evento.is_set():
        print("Processo cancelado antes de salvar o arquivo.")
        return

    wb.save(caminho_arquivo_novo)
    connection.close()

    tempo_total = time.time() - inicio
    if tempo_total > 60:
        minutos = int(tempo_total // 60)
        segundos = tempo_total % 60
        print(f"Tempo total para preencher planilha: {minutos} minutos e {segundos:.0f} segundos\n")
    else:
        print(f"Tempo total para preencher planilha: {tempo_total:.0f} segundos\n")