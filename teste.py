import warnings
from openpyxl import load_workbook

# Suprime os avisos do tipo UserWarning
warnings.filterwarnings("ignore", category=UserWarning)

# Carrega o arquivo
wb = load_workbook("Cadastros Auto Nextt - Skina.xlsm", data_only=True)
ws = wb["Cadastro de Pedidos"]

# Lê e converte a data
valor = ws["D8"].value
print(valor)  # Deve mostrar algo como: 2025-04-29 00:00:00
